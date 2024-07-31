import subprocess
import logging

# 检查
_result = subprocess.run("pip show openpyxl", shell=True, capture_output=True)
if _result.returncode != 0:
    logging.error("This tool is require 'openpyxl' package, please run: pip install openpyxl")
    exit(1)


import sys
import json
import os
import openpyxl
from openpyxl import Workbook, cell
from openpyxl.styles.colors import Color, RGB
from openpyxl.styles.borders import Border, Side


def dump_xlsx_to_json(xlsx_file_path: str, output_json_path: str):
    workbook: Workbook = openpyxl.load_workbook(xlsx_file_path)

    data: dict[str, dict] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Sheet
        sheet_data: list[list[dict[str, None]]] = []
        for row in range(sheet.max_row):
            row_data: list[dict] = []
            for column in range(sheet.max_column):
                cell = sheet.cell(row=row + 1, column=column + 1)
                row_data.append(
                    {
                        "value": cell.value,
                        "fill": _to_dict(cell.fill),
                        "border": _to_dict(cell.border),
                        # TODO: 存储更多的单元格属性
                    }
                )
            sheet_data.append(row_data)

        data[sheet_name] = {
            "data": sheet_data,
            "sheet_format": _to_dict(sheet.sheet_format),
            # TODO: 存储更多的Sheet属性
        }

    workbook.close()
    f = open(output_json_path, "w", encoding="utf8")
    json.dump(obj=data, fp=f, indent="\t")
    f.close()
    return True


def override_xlsx_worksheets_from_json(json_file_path: str, output_xlsx_file_path: str):
    f = open(json_file_path, "r", encoding="utf8")
    data: dict[str, dict] = json.load(f)

    if len(data) <= 0:
        logging.error("The input json has not data.")
        exit(1)

    workbook: Workbook
    default_sheetnames: list[str] = []
    if os.path.exists(output_xlsx_file_path):
        workbook = openpyxl.load_workbook(output_xlsx_file_path)
    else:
        workbook = Workbook()
        default_sheetnames = workbook.sheetnames

    for sheet_name in data:
        if sheet_name in default_sheetnames:
            default_sheetnames.remove(sheet_name)

        # 覆盖
        if not sheet_name in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        if "sheet_format" in data[sheet_name]:
            sheet_format: dict = data[sheet_name]["sheet_format"]
            for k in sheet_format:
                setattr(sheet.sheet_format, k, sheet_format[k])
        # TODO 其他表格属性

        # 数据行
        sheet_data: list[list[dict[str, None]]] = data[sheet_name]["data"]
        for row in range(len(sheet_data)):
            for column in range(len(sheet_data[row])):
                cell_data: dict[str, None] = sheet_data[row][column]
                sheet.cell(row=row + 1, column=column + 1, value=cell_data["value"])
                cell = sheet.cell(row=row + 1, column=column + 1)
                if "fill" in cell_data:
                    _set_cell_fill(cell, cell_data["fill"])
                if "border" in cell_data:
                    _set_cell_border(cell, cell_data["border"])
                # TODO: 其他单元格属性(先检查是否有对应属性)

            # 清除右侧的多余格
            for column in range(len(sheet_data[row]), sheet.max_column):
                cell = sheet.cell(row=row + 1, column=column + 1)
                if not cell.value is None:
                    cell.value = ""

    for sheet_name in default_sheetnames:
        if sheet_name in workbook:
            del workbook[sheet_name]

    workbook.save(output_xlsx_file_path)
    workbook.close()


def _to_dict(obj: None) -> dict:
    if not hasattr(obj, "__dict__"):
        return obj
    ret = {}
    for k in vars(obj):
        if k.startswith("_"):
            continue
        v = getattr(obj, k)
        if hasattr(v, "__dict__"):
            v = _to_dict(v)
        ret[k] = v
    return ret


def _set_cell_fill(cell: cell.Cell, dict: dict) -> None:
    import openpyxl.styles.fills as fills

    PatternFill_attrs = ["patternType", "fgColor", "bgColor"]
    GradientFill_attrs = ["type", "degree", "left", "right", "top", "bottom", "stop"]

    is_meet = lambda attrs: len([attr for attr in attrs if attr in dict]) == len(attrs)

    if is_meet(PatternFill_attrs):
        cell.fill = fills.PatternFill(
            patternType=dict["patternType"],
            fgColor=_get_color_from_dict(dict["fgColor"]),
            bgColor=_get_color_from_dict(dict["bgColor"]),
        )
    elif is_meet(GradientFill_attrs):
        sl = []
        for dict in dict["stop"]:
            stop = fills.Stop(color=_get_color_from_dict(dict=["color"]), position=dict["position"])
            sl.append(stop)
        cell.fill = fills.GradientFill(
            type=dict["type"],
            degree=dict["degree"],
            left=dict["left"],
            right=dict["right"],
            top=dict["top"],
            bottom=dict["bottom"],
            stop=sl,
        )
    else:
        logging("Parse error: is not a valid dict.", dict)


def _set_cell_border(cell: cell.Cell, dict: dict) -> None:
    for k in ["start", "end", "left", "right", "top", "bottom", "diagonal", "vertical", "horizontal"]:
        if k not in dict:
            dict[k] = None
        else:
            side_dict = dict[k]
            if not "style" in side_dict:
                side_dict["style"] = None
            if not "color" in side_dict:
                side_dict["color"] = None
            dict[k] = Side(style=side_dict["style"], color=_get_color_from_dict(side_dict["color"]))

    for k in ["diagonalUp", "diagonalDown"]:
        if k not in dict:
            dict[k] = False

    if "outline" not in dict:
        dict["outline"] = True

    cell.border = Border(
        start=dict["start"],
        end=dict["end"],
        left=dict["left"],
        right=dict["right"],
        top=dict["top"],
        bottom=dict["bottom"],
        diagonal=dict["diagonal"],
        vertical=dict["vertical"],
        horizontal=dict["horizontal"],
        diagonalUp=dict["diagonalUp"],
        diagonalDown=dict["diagonalDown"],
        outline=dict["outline"],
    )


def _get_color_from_dict(dict: dict) -> Color:
    if "index" in dict:
        dict["indexed"] = dict["index"]
    if not "tint" in dict:
        dict["tint"] = 0.0
    if "indexed" in dict:
        return Color(indexed=dict["indexed"], tint=dict["tint"])
    elif "theme" in dict:
        return Color(theme=dict["theme"], tint=dict["tint"])
    elif "auto" in dict:
        return Color(auto=dict["auto"])
    else:
        return Color(rgb=dict["rgb"])


def main(argv):
    for arg in argv:
        if arg in ["-h", "--help"]:
            print("Valid command:")
            print("\tpython xlsx_json.py --dump_json path/to/excel.xlsx path/to/output.json")
            print("\tpython xlsx_json.py --override_xlsx path/to/data.json path/to/output.xlsx")
            exit(0)

    invalid_hint = "Invalid arguments for 'xlsx_json' tool. Please type '-h' or '--help' for more details."
    if len(argv) != 3:

        logging.error(invalid_hint)
        exit(1)

    if argv[0] == "--dump_json":
        dump_xlsx_to_json(argv[1], argv[2])
    elif argv[0] == "--override_xlsx":
        override_xlsx_worksheets_from_json(argv[1], argv[2])
    else:

        logging.error(invalid_hint)
        exit(1)


if __name__ == "__main__":
    main(sys.argv[1:])
